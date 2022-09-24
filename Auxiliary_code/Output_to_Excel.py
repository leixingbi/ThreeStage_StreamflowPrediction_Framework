"""
functions:

"""

import sys
import xlwings as xw
import openpyxl
import numpy as np
import os
import pathlib

sys.path.append("../")

excel_sheet_index = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]


def to_excel(data, method, file_name):
    """
    :param data: matrix or time series
    :param method: method name (str)
    :return:
    """

    path = pathlib.Path(".\\Result\\result_multi_process.xlsx")
    if path.exists():
        pass
        # print("The Excel file for recording the result data exists")
    else:
        wb = openpyxl.Workbook()
        bb = ".\\Result\\" + "result_multi_process" + ".xlsx"
        print(bb)
        wb.save(bb)

    data = np.array(data)
    if data.ndim == 2:
        imfs_num, series_num = data.shape
        if imfs_num > series_num:
            data = data.T
    if file_name == "single_process":
        wb = openpyxl.load_workbook(r".\Result\result.xlsx")
    if file_name == "multi_process":
        wb = openpyxl.load_workbook(r".\Result\result_multi_process.xlsx")
    sheet_name = wb.sheetnames
    status = 0
    for i in sheet_name:
        if i == method:
            status = 1
    if status == 0:
        wb.create_sheet(title=method, index=0)
    if file_name == "single_process":
        wb.save(r".\Result\result.xlsx")
    if file_name == "multi_process":
        wb.save(r".\Result\result_multi_process.xlsx")
    wb.close()

    app = xw.App(visible=True, add_book=False)
    if file_name == "single_process":
        wb = app.books.open(r".\Result\result.xlsx")
    if file_name == "multi_process":
        wb = app.books.open(r".\Result\result_multi_process.xlsx")
    sheet = wb.sheets[method]
    if data.ndim == 1:  # One-dimensional time series
        for i in range(len(data)):
            sheet.range("A" + str(i + 1)).value = data[i]
    if data.ndim == 2:  # Two-dimensional time series
        imfs_num, series_num = data.T.shape
        for i in range(imfs_num):
            for j in range(series_num):
                sheet.range(excel_sheet_index[j] + str(i + 1)).value = data[j, i]
    wb.save()
    wb.close()
    app.quit()

